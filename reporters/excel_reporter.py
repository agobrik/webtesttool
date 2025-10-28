"""
Excel Report Generator
Detailed Excel reports with multiple worksheets, charts, and formatting
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.chart import (
    PieChart, BarChart, Reference
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from loguru import logger

from core.exceptions import ReportGenerationError


class ExcelReporter:
    """
    Professional Excel report generator

    Features:
    - Multiple worksheets (Summary, Vulnerabilities, Performance, etc.)
    - Charts and graphs
    - Conditional formatting
    - Styled tables
    - Auto-sizing columns
    """

    # Color scheme (RGB hex)
    COLORS = {
        'critical': 'DC143C',  # Crimson
        'high': 'FF6347',      # Tomato
        'medium': 'FFA500',    # Orange
        'low': 'FFD700',       # Gold
        'info': '87CEEB',      # SkyBlue
        'pass': '32CD32',      # LimeGreen
        'header': '2E4053',    # Dark blue-grey
        'subheader': '5D6D7E', # Medium grey
    }

    def __init__(self, output_dir: str = "reports", include_charts: bool = True):
        """
        Initialize Excel reporter

        Args:
            output_dir: Directory for Excel reports
            include_charts: Include charts and graphs
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.include_charts = include_charts

    def generate(
        self,
        scan_data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        Generate Excel report

        Args:
            scan_data: Scan results data
            filename: Output filename (optional)

        Returns:
            Path to generated Excel file
        """
        try:
            # Generate filename
            if not filename:
                from urllib.parse import urlparse

                # Extract site name from URL
                target_url = scan_data.get('target', 'unknown-site')
                try:
                    parsed_url = urlparse(target_url)
                    site_name = parsed_url.netloc.replace(':', '-').replace('.', '-')
                    if not site_name:
                        site_name = "unknown-site"
                except:
                    site_name = "unknown-site"

                # Determine test type from vulnerabilities
                vulns = scan_data.get('vulnerabilities', [])
                if vulns:
                    test_types = set()
                    for v in vulns[:10]:  # Sample first 10
                        vtype = v.get('type', 'test')
                        test_types.add(vtype)
                    test_name = '-'.join(sorted(list(test_types))[:3])
                else:
                    test_name = "fullscan"

                # Create timestamp
                timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

                # Create filename: site-test-datetime.xlsx
                filename = f"{site_name}-{test_name}-{timestamp}.xlsx"

            filepath = self.output_dir / filename

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create worksheets
            self._create_summary_sheet(wb, scan_data)

            if 'vulnerabilities' in scan_data:
                self._create_vulnerabilities_sheet(wb, scan_data['vulnerabilities'])

            if 'security' in scan_data:
                self._create_security_sheet(wb, scan_data['security'])

            if 'performance' in scan_data:
                self._create_performance_sheet(wb, scan_data['performance'])

            if 'seo' in scan_data:
                self._create_seo_sheet(wb, scan_data['seo'])

            # Save workbook
            wb.save(filepath)

            logger.info(f"Excel report generated: {filepath}")
            return str(filepath)

        except Exception as e:
            raise ReportGenerationError(
                f"Failed to generate Excel report: {str(e)}",
                details={'filename': filename},
                original_error=e
            )

    def _create_summary_sheet(self, wb: Workbook, scan_data: Dict):
        """Create summary worksheet"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Security Scan Report - Summary"
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30

        # Scan information
        row = 3
        info_data = [
            ['Target', scan_data.get('target', 'Unknown')],
            ['Scan Date', scan_data.get('date', 'N/A')],
            ['Duration', scan_data.get('duration', 'N/A')],
            ['Pages Scanned', scan_data.get('summary', {}).get('pages_scanned', 0)],
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        # Summary statistics
        row += 2
        ws[f'A{row}'] = "Issues Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Header row
        headers = ['Severity', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Data rows
        summary = scan_data.get('summary', {})
        total_issues = summary.get('total_issues', 0) or 1  # Avoid division by zero

        severity_data = [
            ('Critical', summary.get('critical', 0), self.COLORS['critical']),
            ('High', summary.get('high', 0), self.COLORS['high']),
            ('Medium', summary.get('medium', 0), self.COLORS['medium']),
            ('Low', summary.get('low', 0), self.COLORS['low']),
        ]

        for severity, count, color in severity_data:
            ws[f'A{row}'] = severity
            ws[f'A{row}'].fill = PatternFill(start_color=color, fill_type='solid')
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count / total_issues * 100):.1f}%"

            # Center align
            for col in range(1, 4):
                ws.cell(row, col).alignment = Alignment(horizontal='center')

            row += 1

        # Add pie chart
        if self.include_charts and total_issues > 0:
            chart = PieChart()
            chart.title = "Issues by Severity"
            chart.height = 10
            chart.width = 15

            # Data reference
            data = Reference(ws, min_col=2, min_row=row-4, max_row=row-1)
            labels = Reference(ws, min_col=1, min_row=row-4, max_row=row-1)

            chart.add_data(data)
            chart.set_categories(labels)

            ws.add_chart(chart, f"E{row-6}")

        # Auto-size columns
        self._autosize_columns(ws)

    def _create_vulnerabilities_sheet(self, wb: Workbook, vulnerabilities: List[Dict]):
        """Create vulnerabilities worksheet"""
        ws = wb.create_sheet("Vulnerabilities")

        # Title
        ws['A1'] = "Vulnerabilities"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:F1')

        # Headers
        headers = ['ID', 'Severity', 'Type', 'Description', 'Location', 'Remediation']
        row = 2

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        row = 3
        for i, vuln in enumerate(vulnerabilities, 1):
            severity = vuln.get('severity', 'Low')
            color = self.COLORS.get(severity.lower(), 'FFFFFF')

            ws[f'A{row}'] = i
            ws[f'B{row}'] = severity
            ws[f'B{row}'].fill = PatternFill(start_color=color, fill_type='solid')
            ws[f'C{row}'] = vuln.get('type', 'Unknown')
            ws[f'D{row}'] = vuln.get('description', '')
            ws[f'E{row}'] = vuln.get('location', '')
            ws[f'F{row}'] = vuln.get('remediation', '')

            # Wrap text
            for col in range(1, 7):
                ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Apply borders
        self._apply_table_borders(ws, 2, 1, row-1, 6)

        # Auto-size columns
        self._autosize_columns(ws)

        # Set column widths manually for better readability
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 40

    def _create_security_sheet(self, wb: Workbook, security_data: Dict):
        """Create security overview worksheet"""
        ws = wb.create_sheet("Security")

        # Title
        ws['A1'] = "Security Overview"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:D1')

        # Security score
        row = 3
        ws[f'A{row}'] = "Overall Security Score"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{security_data.get('score', 0)}/100"
        ws[f'B{row}'].font = Font(size=14, bold=True)

        # Tests performed
        row += 3
        ws[f'A{row}'] = "Tests Performed"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Table headers
        headers = ['Test Name', 'Status', 'Issues Found', 'Pass Rate']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Test data
        tests = security_data.get('tests_performed', [])
        for test in tests:
            ws[f'A{row}'] = test.get('name', '')
            ws[f'B{row}'] = test.get('status', '')
            ws[f'C{row}'] = test.get('issues', 0)
            ws[f'D{row}'] = f"{test.get('pass_rate', 0):.1f}%"

            # Color code status
            status = test.get('status', '').lower()
            if status == 'passed':
                ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['pass'], fill_type='solid')
            elif status == 'failed':
                ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['critical'], fill_type='solid')

            row += 1

        # Auto-size columns
        self._autosize_columns(ws)

    def _create_performance_sheet(self, wb: Workbook, performance_data: Dict):
        """Create performance worksheet"""
        ws = wb.create_sheet("Performance")

        # Title
        ws['A1'] = "Performance Analysis"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:D1')

        # Metrics
        row = 3
        ws[f'A{row}'] = "Performance Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Table headers
        headers = ['Metric', 'Value', 'Threshold', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Metrics data
        metrics = performance_data.get('metrics', {})
        metrics_list = [
            ('Average Response Time', f"{metrics.get('avg_response_time', 0)}ms", '< 1000ms'),
            ('Page Load Time', f"{metrics.get('page_load_time', 0)}ms", '< 3000ms'),
            ('Time to First Byte', f"{metrics.get('ttfb', 0)}ms", '< 600ms'),
            ('Total Requests', str(metrics.get('total_requests', 0)), 'N/A'),
        ]

        for metric_name, value, threshold in metrics_list:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            ws[f'C{row}'] = threshold
            ws[f'D{row}'] = 'Good'  # TODO: Calculate based on threshold

            row += 1

        # Auto-size columns
        self._autosize_columns(ws)

    def _create_seo_sheet(self, wb: Workbook, seo_data: Dict):
        """Create SEO worksheet"""
        ws = wb.create_sheet("SEO")

        # Title
        ws['A1'] = "SEO Analysis"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
        ws.merge_cells('A1:C1')

        # SEO checks
        row = 3
        ws[f'A{row}'] = "SEO Checks"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Headers
        headers = ['Check', 'Status', 'Issues']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # SEO checks data
        checks = seo_data.get('checks', [])
        for check in checks:
            ws[f'A{row}'] = check.get('name', '')
            ws[f'B{row}'] = check.get('status', '')
            ws[f'C{row}'] = check.get('issues', 0)

            # Color code status
            if check.get('status') == 'Pass':
                ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['pass'], fill_type='solid')
            else:
                ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['medium'], fill_type='solid')

            row += 1

        # Auto-size columns
        self._autosize_columns(ws)

    def _apply_table_borders(
        self,
        ws: Worksheet,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int
    ):
        """Apply borders to table"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row, col).border = thin_border

    def _autosize_columns(self, ws: Worksheet):
        """Auto-size columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width


# Convenience function

def generate_excel_report(
    scan_data: Dict[str, Any],
    output_dir: str = "reports",
    filename: Optional[str] = None
) -> str:
    """
    Generate Excel report (convenience function)

    Args:
        scan_data: Scan results data
        output_dir: Output directory
        filename: Output filename

    Returns:
        Path to generated Excel file
    """
    reporter = ExcelReporter(output_dir=output_dir)
    return reporter.generate(scan_data, filename)
